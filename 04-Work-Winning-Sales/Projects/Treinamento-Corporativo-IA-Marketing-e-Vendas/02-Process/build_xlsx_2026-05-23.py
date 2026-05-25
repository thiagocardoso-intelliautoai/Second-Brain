# -*- coding: utf-8 -*-
"""
Script historico do prototipo .xlsx v3, hoje descartado como entregavel.
Baseado em 02-Process/2026-05-23-Escopo-Final-Consolidado.md
e 02-Process/2026-05-23-Plano-Adaptacao-Sheets.md.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"D:\1A-Projetos-Pessoais\Second Brain\04-Work-Winning-Sales\Projects\Treinamento-Corporativo-IA-Marketing-e-Vendas\02-Process\Treinamento-Corporativo-IA-Marketing-e-Vendas-v3.xlsx"

# Historical script only. The .xlsx v3 path was discarded as the current
# deliverable; the active output is the Google Sheets edited in-place.

# ----------- Estilos -----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BLOCK_TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
SUBHEADER_FONT = Font(bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_SIDE = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# ----------- Definição das abas -----------

TABS = [
    # ============================================================
    # ABA 1 - RESUMO EXECUTIVO
    # ============================================================
    {
        "name": "1. Resumo Executivo",
        "widths": [32, 100],
        "blocks": [
            {
                "title": None,
                "headers": ["Campo", "Conteúdo"],
                "rows": [
                    ["Promessa central", "Sair de 'uso pontual de IA' para rotinas implantadas em 4 semanas, com ganho de tempo mensurável, qualidade padronizada e outputs reais revisados."],
                    ["Posicionamento", "Consultoria boutique brasileira, método replicável documentado, voz consultiva nativa em português, para B2B de médio/grande porte."],
                    ["Princípio pedagógico", "Menos aula inspiracional, mais execução aplicada ao negócio do cliente. Culmina em 18 SKILL.md operacionais + plano de adoção de 30 dias com sponsor formalmente engajado."],
                    ["Público-alvo (cohort cliente)", "Times de Marketing B2B (Gestor/Head/CMO, Estrategista de Conteúdo, Especialista de Demanda) + Times de Vendas B2B (Gestor Comercial, SDR/BDR, AE/Closer). Cohort 8-20 pessoas."],
                    ["Sponsor (decisor cliente)", "CRO, CMO, VP Vendas, Head Marketing ou CEO em empresa menor. Presença obrigatória nos últimos 15 min da Sessão 4."],
                    ["SKU CORE (per-engagement)", "4 sessões × 1h, trilha híbrida (S1/S4 mistas, S2 marketing, S3 vendas), biblioteca async inclusa (18 SKILL.md + vídeos + templates), pricing R$ 30-80k por cliente."],
                    ["SKU PREMIUM (add-on opcional)", "4 semanas pós-treinamento (semanas 5-8), 2 checkpoints quinzenais + 1 review executivo, relatório de impacto 30 dias com métricas antes/depois. +30-40% do CORE."],
                    ["Entregáveis CORE", "8 entregáveis: maturity score, diagnóstico, mapa priorizado, biblioteca 18 SKILL.md, outputs revisados, rubrica de qualidade, doc governança, plano de adoção 30d. Detalhes na Aba 4."],
                    ["Entregáveis PREMIUM", "3 adicionais: checkpoints documentados, relatório de impacto 30d executivo, apresentação final ao sponsor. Detalhes na Aba 7."],
                    ["Diferenciadores", "Português nativo, contexto B2B brasileiro, método replicável documentado, 18 SKILL.md operacionais prontos, foco aplicado (vs estratégico de C-suite), Maturity Model próprio, Winning AI como acelerador opcional. Ver Aba 10."],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 2 - CRONOGRAMA E ARQUITETURA
    # ============================================================
    {
        "name": "2. Cronograma e Arquitetura",
        "widths": [38, 10, 12, 18, 50, 28, 10],
        "blocks": [
            {
                "title": None,
                "headers": ["Macro entregável", "Semana", "Duração", "Tipo", "Entregável", "Responsável cliente", "SKU"],
                "rows": [
                    ["Diagnóstico pre-work + maturity score", "0", "0,5h/pessoa", "Preparação", "Formulário pre-work + score automático (1-5) em 5 dimensões + recomendação de SKU.", "Sponsor + líderes de marketing/vendas", "CORE"],
                    ["Sessão 1 — Como criar skill por ROI + Diagnóstico", "1", "1h", "Treinamento/Workshop", "Top 5 casos de uso priorizados + 1 skill genérica criada ao vivo.", "Turma completa", "CORE"],
                    ["Tarefa S1 — Mapa de workflows", "1", "0,5h", "Exercício", "Planilha com 10-15 oportunidades de IA na função.", "Cada participante", "CORE"],
                    ["Tarefa S1 — Contexto mínimo", "1", "0,5h", "Exercício", "Pasta de briefing do negócio (ICP, oferta, playbook, tom de voz, métricas).", "Marketing Ops / RevOps / Gestores", "CORE"],
                    ["Sessão 2 — IA aplicada em Marketing", "2", "1h", "Treinamento/Workshop", "3 skills âncora rodadas ao vivo (1 por persona).", "Time de marketing", "CORE"],
                    ["Tarefa S2 — Ativo de marketing", "2", "1h", "Exercício", "1 ativo real (landing, email, post, anúncio, briefing) rodado com IA.", "Cada participante de marketing", "CORE"],
                    ["Sessão 3 — IA aplicada em Vendas", "3", "1h", "Treinamento/Workshop", "3 skills âncora rodadas ao vivo (1 por persona).", "Time de vendas", "CORE"],
                    ["Tarefa S3 — Fluxo de vendas", "3", "1h", "Exercício", "1 fluxo real (mensagem, prep, follow-up, MAP) rodado com IA + antes/depois.", "Cada participante de vendas", "CORE"],
                    ["Sessão 4 — Governança + Adoção + Sponsor", "4", "1h", "Workshop", "Plano de adoção 30 dias por pessoa + governança + rubrica + sponsor approval.", "Turma + sponsor (últimos 15 min)", "CORE"],
                    ["Checkpoint 1", "5", "1h", "Acompanhamento", "Pilot review, ajustes nos workflows, desbloqueio de impedimentos.", "Champions + consultor Winning", "PREMIUM"],
                    ["Checkpoint 2", "7", "1h", "Acompanhamento", "Métricas parciais, expansão para outros membros do time.", "Champions + consultor Winning", "PREMIUM"],
                    ["Review Executivo", "8", "1h", "Apresentação", "Relatório de impacto 30 dias com métricas antes/depois (uso, qualidade, tempo, impacto comercial).", "Sponsor + champions", "PREMIUM"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 3 - SESSÕES DETALHADAS
    # ============================================================
    {
        "name": "3. Sessões Detalhadas",
        "widths": [6, 38, 22, 50, 50, 50, 12, 40, 35],
        "blocks": [
            {
                "title": None,
                "headers": ["Sessão", "Tema", "Público", "Objetivo", "Conteúdo (15 min)", "Workshop (30 min)", "Pílulas (5 min)", "Combinados + Tarefa (10 min)", "Output esperado"],
                "rows": [
                    [
                        "1",
                        "Como criar skill por ROI + Diagnóstico",
                        "Turma completa (marketing + vendas + gestão)",
                        "Ensinar metaframework: identificar onde IA gera retorno real e transformar em skill operacional.",
                        "Introdução (5'): por que sair de 'uso pontual' para 'skill operacional'. Mapeamento de processos (parte do workshop). Matriz de priorização por ROI (frequência × tempo × dados × clareza × risco). Fit check IA. Pesquisa de soluções existentes (GitHub, OpenClaudia, GTM Agents, prompts públicos).",
                        "Turma cria 1 skill genérica ao vivo (marketing OU vendas). Cada participante mapeia 5-8 processos repetitivos da função e prioriza. Consultor revisa 2-3 mapas em voz alta.",
                        "Q&A rápido, troca de descobertas sobre processos repetitivos.",
                        "Tarefa: mapa de workflows (10-15 oportunidades) + contexto mínimo (pasta de briefing). Combinados: próxima sessão e quem traz o quê.",
                        "Top 5 casos de uso priorizados por impacto e facilidade.",
                    ],
                    [
                        "2",
                        "IA Aplicada em Marketing",
                        "Só time de marketing (Gestor, Estrategista de Conteúdo, Demanda/Growth)",
                        "Rodar 1 skill âncora por persona com ativo real do cliente.",
                        "Apresentação curta das 3 skills âncora (1 por persona): Gestor → Diagnóstico Semanal de Pipeline; Estrategista → Brief Estratégico de Conteúdo; Demanda → Inteligência de Conta + Outreach Brief.",
                        "Cada participante roda a skill da persona dele em ATIVO REAL da semana de trabalho. Consultor circula, revisa 2-3 outputs ao vivo aplicando a rubrica de qualidade (contexto real? próximo passo? não inventa dado? revisável?).",
                        "Q&A sobre context engineering e gaps de contexto.",
                        "Tarefa: entregar 1 ativo de marketing rodado com IA até 4ª-feira no canal #treinamento-ia. Trazer output revisado pra S3 (contato cruzado marketing-vendas).",
                        "1 ativo de marketing pronto para revisão/publicação assistida.",
                    ],
                    [
                        "3",
                        "IA Aplicada em Vendas",
                        "Só time de vendas (Gestor Comercial, SDR/BDR, AE/Closer)",
                        "Rodar 1 skill âncora por persona com caso real do pipeline.",
                        "Apresentação curta das 3 skills âncora (1 por persona): Gestor → Radar Semanal de Forecast e Deal Risk; SDR/BDR → Copiloto de Vendas (6 ângulos); AE → Raio-X de Deal MEDDICC + MAP.",
                        "Cada participante usa um deal/lead/conta real do CRM. Consultor revisa 2-3 outputs ao vivo. Registrar antes/depois (tempo gasto, qualidade percebida).",
                        "Q&A sobre data quality e context engineering em vendas.",
                        "Tarefa: entregar 1 fluxo de vendas com IA (mensagem, plano de conta, prep, follow-up, update CRM) com evidência real. Trazer antes/depois pra S4.",
                        "1 output de vendas usado ou pronto para uso, com qualidade revisada.",
                    ],
                    [
                        "4",
                        "Governança, Adoção e Sponsor Approval",
                        "Turma completa (marketing + vendas + gestão), sponsor nos últimos 15 min",
                        "Transformar aprendizado das sessões 1-3 em rotina operacional permanente, com aprovação executiva visível.",
                        "Conteúdo enxuto: rubrica de qualidade explícita + governança das skills (quem mantém, aprova, onde ficam, como time descobre).",
                        "80% workshop: Bloco 1 — Pilot review (15'). Bloco 2 — Plano de adoção 30 dias por pessoa em template (25'). Bloco 3 — Decisões de governança (10'). Bloco 4 — Rubrica de qualidade aplicada (10'). Bloco 5 — Sponsor approval formal (5', com sponsor presente).",
                        "(Não há pílulas tradicionais — sessão é workshop puro.)",
                        "Combinados finais: cadência de checkpoints, quem reporta progresso ao sponsor em 15d e 30d. Sponsor approve público.",
                        "Plano de adoção 30 dias por pessoa + sponsor approval documentado + governança definida.",
                    ],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 4 - EXERCÍCIOS E REVISÃO
    # ============================================================
    {
        "name": "4. Exercícios e Revisão",
        "widths": [38, 22, 60, 60, 18],
        "blocks": [
            {
                "title": None,
                "headers": ["Exercício", "Quando", "Output a entregar", "Critério de revisão", "Obrigatório?"],
                "rows": [
                    ["Mapa de workflows", "Após S1", "Planilha com 10-15 oportunidades de IA na função.", "Clareza da tarefa, frequência, dados disponíveis, output avaliável.", "Sim"],
                    ["Contexto mínimo", "Após S1", "Pasta/briefing de contexto do negócio (ICP, oferta, playbook, tom de voz, campanhas, funil, métricas).", "Específico, atualizado, aplicável. Sem contexto, IA vira output genérico.", "Sim"],
                    ["Ativo de marketing com IA", "Após S2", "Landing, email, post, anúncio, roteiro de webinar ou briefing.", "Proposta clara, aderência ao ICP, critério de conversão.", "Sim (marketing)"],
                    ["Fluxo de vendas com IA", "Após S3", "Mensagem, plano de conta, proposta, resumo de call ou update de CRM.", "Evidência real, personalização, próximo passo claro.", "Sim (vendas)"],
                    ["Plano de Adoção 30 dias (por pessoa)", "Em S4 (workshop)", "Template preenchido: skill + frequência + workflow específico + métrica + checkpoint.", "Específico, mensurável, com owner e data.", "Sim"],
                    ["Antes/depois — baseline vs S4", "Em S4", "Registro de tempo gasto e qualidade percebida por workflow piloto.", "Comparação quantificada honesta.", "Sim"],
                    ["Relatório de impacto 30 dias", "Semana 8 (Premium)", "Documento executivo com métricas em 4 categorias: uso, qualidade, tempo, impacto comercial.", "Dados antes/depois suficientes para decidir escala.", "Sim (PREMIUM)"],
                ],
            },
            {
                "title": "Diferenças vs versão Rafinha",
                "headers": ["Mudança", "Detalhe"],
                "rows": [
                    ["REMOVIDO", "Canvas de Skill como exercício separado após S4 (redundante com plano de adoção; canvas já é trabalho prático do Módulo 1 em S1)."],
                    ["MODIFICADO", "Relatório de ganho deixou de ser 'opcional premium' para 'obrigatório no PREMIUM'. Sem relatório, cliente não justifica continuidade."],
                    ["ADICIONADO", "Exercício formal de antes/depois (baseline pre-work vs S4) para gerar dados de mensuração."],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 5 - BIBLIOTECA DE SKILLS (18 SKILLS)
    # ============================================================
    {
        "name": "5. Biblioteca de Skills",
        "widths": [4, 11, 30, 60, 14, 25, 14, 14, 14],
        "blocks": [
            {
                "title": None,
                "headers": ["#", "Área", "Persona", "Skill", "Tier", "Tipo de uso", "SKILL.md", "Vídeo curto", "Prioridade build"],
                "rows": [
                    [1, "Vendas", "Gestor Comercial B2B", "Radar Semanal de Forecast e Deal Risk", "Tier 1 (S3)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [2, "Vendas", "Gestor Comercial B2B", "Coach de Performance Comercial por Rep", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [3, "Vendas", "Gestor Comercial B2B", "Radar de Cobertura de Pipeline e Alocação Comercial", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [4, "Vendas", "Especialista de Pipeline (SDR/BDR)", "Copiloto de Vendas (6 ângulos de abordagem)", "Tier 1 (S3)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [5, "Vendas", "Especialista de Pipeline (SDR/BDR)", "Follow-up Estratégico (6 frentes alternativas)", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [6, "Vendas", "Especialista de Pipeline (SDR/BDR)", "Qualificador e Priorizador de Leads/Contas", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [7, "Vendas", "Executivo de Contas (AE)", "Raio-X de Deal MEDDICC + Mutual Action Plan", "Tier 1 (S3)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [8, "Vendas", "Executivo de Contas (AE)", "Discovery-to-Value Case Builder", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [9, "Vendas", "Executivo de Contas (AE)", "Plano de Conta + Expansão por Sinais", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [10, "Marketing", "Gestor de Marketing B2B", "Diagnóstico Semanal de Pipeline e Forecast de Marketing", "Tier 1 (S2)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [11, "Marketing", "Gestor de Marketing B2B", "Priorizador de Campanhas, Canais e Orçamento", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [12, "Marketing", "Gestor de Marketing B2B", "Loop de Alinhamento Marketing-Vendas-Produto", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [13, "Marketing", "Estrategista de Conteúdo e Posicionamento", "Brief Estratégico de Conteúdo por Intenção, POV e Prova", "Tier 1 (S2)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [14, "Marketing", "Estrategista de Conteúdo e Posicionamento", "Minerador de Voz do Cliente e Sinais de Mercado", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [15, "Marketing", "Estrategista de Conteúdo e Posicionamento", "Message House de Posicionamento Defensável", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [16, "Marketing", "Especialista de Demanda e Growth", "Inteligência de Conta + Outreach Brief", "Tier 1 (S2)", "Workshop ao vivo + async", "A produzir", "A produzir", "Alta"],
                    [17, "Marketing", "Especialista de Demanda e Growth", "Diagnóstico de Performance + Rebalanceamento de Campanhas", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                    [18, "Marketing", "Especialista de Demanda e Growth", "Triage de Sinais de Demanda + Handoff MQL→SQL", "Tier 2 (async)", "Só biblioteca async", "A produzir", "A produzir", "Média"],
                ],
            },
            {
                "title": "Recursos complementares (incluídos na biblioteca async)",
                "headers": ["Recurso", "Quantidade", "Status"],
                "rows": [
                    ["Vídeos curtos (5-10 min) por skill", "18 vídeos", "A produzir"],
                    ["Template Plano de Adoção 30 dias", "1", "A produzir"],
                    ["Template Maturity Scorecard (formulário pre-work)", "1", "A produzir"],
                    ["Template Rubrica de Qualidade de Output", "1", "A produzir"],
                    ["Template Canvas de Skill (para criação de novas)", "1", "A produzir"],
                    ["Cheat sheet de Context Engineering", "1", "A produzir"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 6 - PRE-WORK, MATURITY SCORE & DIAGNÓSTICO
    # ============================================================
    {
        "name": "6. Pre-work & Maturity",
        "widths": [28, 60, 38],
        "blocks": [
            {
                "title": "Bloco 1 — Formulário Pre-work (8 dimensões)",
                "headers": ["Dimensão", "Pergunta", "Escala 1-5"],
                "rows": [
                    ["Uso atual de IA", "Quantos % do time usa IA pelo menos 1x/semana em workflow real?", "1: <20% / 5: >80%"],
                    ["Stack tecnológico", "Quais ferramentas de IA já existem (ChatGPT, Claude, Gemini, Copilot, vendor-specific)?", "1: nenhuma / 5: stack alinhado com governance"],
                    ["CRM e dados", "Como está a qualidade dos dados no CRM e a integração com outros sistemas?", "1: bagunçado, sem confiança / 5: dados limpos, integrados"],
                    ["Processo de vendas", "Existe metodologia explícita (MEDDICC, BANT, Challenger)? Playbook documentado?", "1: ad hoc / 5: metodologia clara, seguida pelo time"],
                    ["Marketing ops", "Tem attribution funcionando, lead scoring, marketing automation?", "1: vibe marketing / 5: marketing ops maduro"],
                    ["Sponsor e liderança", "CRO/CMO usa IA visivelmente? Cobra adoção?", "1: não usa, não cobra / 5: lidera adoção ativamente"],
                    ["Apetite de mudança", "Cultura do time é de experimentação ou aversão a risco?", "1: conservador / 5: piloto-friendly"],
                    ["Capacity", "Quem teria 2-4 semanas de bandwidth para fazer o piloto?", "1: ninguém / 5: time disponível"],
                ],
            },
            {
                "title": "Bloco 2 — Classificação do Maturity Score",
                "headers": ["Faixa de score médio", "Classificação", "Recomendação de SKU"],
                "rows": [
                    ["1.0 - 2.0", "Inicial", "Pode rodar CORE, mas piloto pequeno (cohort 8-10) e considerar pré-aquecimento informal antes."],
                    ["2.1 - 3.0", "Em desenvolvimento", "CORE + considerar PREMIUM (cliente precisa de follow-up para não cair em pilot purgatory)."],
                    ["3.1 - 4.0", "Definido", "CORE + PREMIUM padrão. Cliente tem maturidade para extrair valor pleno."],
                    ["4.1 - 5.0", "Gerenciado/Otimizado", "CORE pode ser opcional; foco em PREMIUM, skills customizadas adicionais, train-the-trainer."],
                ],
            },
            {
                "title": "Bloco 3 — Output do diagnóstico (entregue 48h antes da S1)",
                "headers": ["Item entregue", "Descrição"],
                "rows": [
                    ["Maturity Score", "Score médio + classificação + scores por dimensão"],
                    ["Mapa de oportunidades preliminar", "Top 10-15 workflows candidatos a piloto, com ROI estimado"],
                    ["Recomendação de priorização", "Por área (marketing/vendas) e por persona"],
                    ["Recomendação de SKU", "CORE básico, CORE + PREMIUM, ou config customizada"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 7 - ACOMPANHAMENTO PREMIUM
    # ============================================================
    {
        "name": "7. Acompanhamento Premium",
        "widths": [10, 25, 12, 60, 40],
        "blocks": [
            {
                "title": "Estrutura das 4 semanas pós-treinamento (semanas 5-8)",
                "headers": ["Semana", "Atividade", "Duração", "Foco", "Entregável"],
                "rows": [
                    ["5", "Checkpoint 1 (quinzenal)", "1h", "Pilot review: o que cada um está rodando? O que travou? Ajustes nos workflows. Desbloqueio de impedimentos.", "Notas + lista de ações"],
                    ["7", "Checkpoint 2 (quinzenal)", "1h", "Métricas parciais coletadas. Expansão para outros membros do time. Iteração das skills.", "Métricas parciais documentadas + plano de expansão"],
                    ["8", "Review Executivo", "1h", "Apresentação para sponsor: o que mudou em 30 dias? Próximos passos. Decisão de continuidade.", "Relatório de impacto 30 dias + apresentação executiva"],
                ],
            },
            {
                "title": "Métricas medidas no Relatório de Impacto 30 dias",
                "headers": ["Categoria", "Métricas medidas", "Baseline", "30 dias", "Variação"],
                "rows": [
                    ["Uso (adoção)", "% de participantes ativos em 30d; frequência semanal de uso por workflow", "[baseline]", "[medido]", "[delta]"],
                    ["Qualidade", "% de outputs que passam a rubrica; % reutilizados pelo time", "[baseline]", "[medido]", "[delta]"],
                    ["Tempo", "Horas economizadas por workflow; redução de cycle time", "[baseline]", "[medido]", "[delta]"],
                    ["Impacto comercial", "Reply rate, win rate, CPL, conversão, ticket médio, velocidade de pipeline", "[baseline]", "[medido]", "[delta]"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 8 - PRICING & ESTRUTURA COMERCIAL
    # ============================================================
    {
        "name": "8. Pricing",
        "widths": [22, 22, 38, 22],
        "blocks": [
            {
                "title": "Bloco 1 — Pricing CORE (faixa R$ 30-80k por engagement)",
                "headers": ["Calibração", "Tamanho cohort", "Customização incluída", "Pricing"],
                "rows": [
                    ["Básico", "8-10 pessoas", "Customização leve (exemplos genéricos)", "R$ 30-45k"],
                    ["Médio", "10-15 pessoas", "Customização média (slides com brand do cliente, exemplos do contexto)", "R$ 45-65k"],
                    ["Premium config", "15-20 pessoas", "Customização alta (skills adicionais incluídas, sessões extras se necessário)", "R$ 65-80k"],
                ],
            },
            {
                "title": "Bloco 2 — Pricing PREMIUM (add-on opcional, +30-40% do CORE)",
                "headers": ["Calibração CORE", "Premium add-on (+30-40%)"],
                "rows": [
                    ["R$ 30-45k", "+R$ 9-18k"],
                    ["R$ 45-65k", "+R$ 13,5-26k"],
                    ["R$ 65-80k", "+R$ 19,5-32k"],
                ],
            },
            {
                "title": "Bloco 3 — Add-ons comerciais separados",
                "headers": ["Add-on", "Pricing"],
                "rows": [
                    ["Skill customizada (fora dos 18 base)", "R$ 5-25k por skill complexa"],
                    ["Sessão extra (1h)", "R$ 5-10k"],
                    ["Train-the-trainer (formar champion interno)", "30-60% do CORE"],
                    ["Refresher quarterly (workshop 2-4h a cada 3 meses)", "15-25% do CORE"],
                    ["Certificação por participante", "R$ 100-500 por aluno"],
                ],
            },
            {
                "title": "Bloco 4 — Pré-requisitos de fechamento de venda",
                "headers": ["#", "Pré-requisito"],
                "rows": [
                    ["1", "Sponsor identificado com poder de decisão sobre adoção"],
                    ["2", "Uso básico de IA já existente no time (não partimos do zero absoluto)"],
                    ["3", "Capacity declarada: time tem 2-4 semanas de bandwidth para o programa + 4 semanas para piloto"],
                    ["4", "Cohort viável: pelo menos 8 pessoas confirmadas"],
                    ["5", "Ferramentas básicas disponíveis: acesso a ChatGPT, Claude ou Gemini (não exigimos Winning AI)"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 9 - CRITÉRIOS DE QUALIDADE & TEMPLATES
    # ============================================================
    {
        "name": "9. Critérios de Qualidade",
        "widths": [30, 60, 30],
        "blocks": [
            {
                "title": "Bloco 1 — Rubrica de Qualidade de Output (4 perguntas críticas)",
                "headers": ["Pergunta", "Critério de aprovação"],
                "rows": [
                    ["Usa contexto real do negócio?", "Cita ICP específico, oferta, dados reais. Não usa invenções genéricas tipo 'empresas B2B' sem qualificar."],
                    ["Tem próximo passo claro?", "Acionável: 'enviar e-mail X', 'criar tarefa Y no CRM', não apenas 'estudar oportunidade'."],
                    ["Não inventa dado?", "Separa fato (do CRM/dados reais), inferência (extrapolação fundamentada) e hipótese (a validar). Sem fabricar números."],
                    ["É revisável em 2 min?", "Forma estruturada (bullets, tabelas, headers). Não é bloco corrido de texto."],
                ],
            },
            {
                "title": "Bloco 2 — Matriz de Priorização de Workflows (workshop S1)",
                "headers": ["Critério", "Peso", "Pergunta-chave"],
                "rows": [
                    ["Frequência", "Alto", "Quantas vezes por semana esse workflow roda?"],
                    ["Tempo gasto por execução", "Alto", "Quantas horas/minutos consome cada vez?"],
                    ["Dados disponíveis", "Médio", "Há dados estruturados ou contexto pra alimentar IA?"],
                    ["Clareza do output esperado", "Médio", "É possível julgar 'bom output vs ruim'?"],
                    ["Risco se errar", "Baixo", "Erro de IA tem alto impacto operacional/comercial?"],
                ],
            },
            {
                "title": "Bloco 3 — Quality Gates por Skill (para o engenheiro de IA buildar)",
                "headers": ["Gate", "Critério"],
                "rows": [
                    ["Citação de evidência", "Toda recomendação cita a evidência (dado, dataset, fonte) usada"],
                    ["Separação fato/inferência/recomendação", "Output explicita o que é fato, o que é inferência e o que é recomendação acionável"],
                    ["Próximo passo claro", "Termina com owner, próximo passo, métrica impactada"],
                    ["Sinalização de lacunas", "Quando falta dado, a skill sinaliza explicitamente em vez de inventar"],
                ],
            },
            {
                "title": "Bloco 4 — Templates Disponíveis (a produzir)",
                "headers": ["Template", "Formato", "Status"],
                "rows": [
                    ["Plano de Adoção 30 dias", "Excel/Sheets + Notion", "A produzir"],
                    ["Maturity Scorecard (formulário pre-work)", "Google Form + scoring auto", "A produzir"],
                    ["Rubrica de Qualidade de Output", "PDF + Markdown", "A produzir"],
                    ["Canvas de Skill", "Markdown template", "A produzir"],
                    ["Cheat sheet de Context Engineering", "PDF 1-pager", "A produzir"],
                ],
            },
        ],
    },

    # ============================================================
    # ABA 10 - FONTES, MODOS DE FALHA & DIFERENCIADORES
    # ============================================================
    {
        "name": "10. Fontes & Diferenciadores",
        "widths": [40, 70],
        "blocks": [
            {
                "title": "Bloco 1 — Diferenciadores Competitivos (7)",
                "headers": ["#", "Diferenciador"],
                "rows": [
                    ["1", "Português nativo, voz consultiva brasileira (vs traduções de Section/Reforge/Big 4)"],
                    ["2", "Conhecimento de contexto B2B brasileiro (CRM local, sales engagement BR, regulamentação, ciclo de venda)"],
                    ["3", "Método replicável já documentado (vs cada cliente começa do zero)"],
                    ["4", "18 SKILL.md operacionais prontos (vs prompts soltos ou frameworks abstratos)"],
                    ["5", "Foco aplicado (vs estratégico de C-suite — Big 4 ocupa esse espaço)"],
                    ["6", "Maturity Model próprio (alavanca de autoridade quando publicado)"],
                    ["7", "Winning AI como acelerador opcional, não obrigatório (não trava venda)"],
                ],
            },
            {
                "title": "Bloco 2 — 11 Modos de Falha Mitigados pelo método",
                "headers": ["Failure Mode", "Mitigação no método Winning Sales"],
                "rows": [
                    ["Demo Theatre (impressionar sem implantar)", "50%+ do tempo em workshop com output real do aluno + revisão crítica ao vivo"],
                    ["ChatGPT 101 (genérico demais)", "S1 é 'Como criar skill por ROI', não fundamentos de IA"],
                    ["No Metrics (sem dado antes/depois)", "Maturity score baseline + relatório de impacto 30d obrigatório no PREMIUM"],
                    ["One-and-done (sem follow-up)", "Plano de adoção 30 dias produzido em workshop + checkpoints quinzenais no PREMIUM"],
                    ["No Leadership Engagement", "Sponsor formal nos últimos 15 min de S4 + review executivo no PREMIUM"],
                    ["Tool Zoo (muitas ferramentas, pouco workflow)", "Cada skill é 1 workflow, 1 output, 1 critério de qualidade. Não cobrimos todas as ferramentas"],
                    ["No Context Engineering", "Cheat sheet específico + workshops sempre com contexto real do cliente"],
                    ["Ungated Assignments (tarefas sem revisão)", "Toda tarefa de campo é revisada na sessão seguinte"],
                    ["Skill Bloat (skills sem validar uso real)", "Tier 1 (6 skills ao vivo) + Tier 2 (12 async); piloto valida quais viram uso real"],
                    ["Sponsor Disconnect", "Sponsor approval formal em S4 + cadência de report"],
                    ["Pilot Purgatory (piloto não escala)", "PREMIUM endereça semanas 5-8; relatório de impacto cria caso para próxima fase"],
                ],
            },
            {
                "title": "Bloco 3 — Fontes externas (pesquisa de mercado)",
                "headers": ["Fonte", "Tipo / Uso"],
                "rows": [
                    ["McKinsey QuantumBlack Academy [verificar]", "Benchmark de programa enterprise customizado"],
                    ["BCG AI Academy / BCG GenAI Operating Model [verificar]", "Benchmark de governance e capability building"],
                    ["Section.school (AI for Marketers) [verificar]", "Benchmark de boutique cohort americana"],
                    ["Reforge (AI for Growth) [verificar]", "Benchmark de boutique cohort focada em product/growth"],
                    ["Wharton Executive Education (AI for Business) [verificar]", "Benchmark de executive ed"],
                    ["HubSpot Academy + Salesforce Trailhead", "Benchmark de vendor-led training gratuito"],
                    ["Adult learning research", "70-20-10 (Lombardo & Eichinger); Cognitive Apprenticeship (Collins, Brown, Holum); Kirkpatrick four-level; Bloom's revisada; Andragogy (Knowles)"],
                    ["Industry reports", "McKinsey State of AI; BCG GenAI reports; Deloitte Generative AI Dossier; HubSpot State of Sales; Salesforce State of Sales; Gartner CMO Spend Survey"],
                ],
            },
            {
                "title": "Bloco 4 — Documentos internos (Second Brain)",
                "headers": ["Documento", "Caminho"],
                "rows": [
                    ["Escopo Final Consolidado", "02-Process/2026-05-23-Escopo-Final-Consolidado.md"],
                    ["Pesquisa de Mercado", "02-Process/2026-05-23-Pesquisa-Mercado-Treinamento-IA-Corporativo.md"],
                    ["Plano de Alteração (cross-análise)", "02-Process/2026-05-23-Plano-de-Alteracao-Escopo.md"],
                    ["Plano de Adaptação do Sheets", "02-Process/2026-05-23-Plano-Adaptacao-Sheets.md"],
                    ["Escopo Thiago detalhado", "01-Inputs/Aprimoria do Treinamento.md"],
                    ["Escopo Rafinha original (snapshot)", "01-Inputs/Escopo-em-Producao-Google-Sheets.md"],
                    ["Project Brief", "00-Project-Brief.md"],
                    ["Project Cockpit", "PROJECT.md"],
                ],
            },
        ],
    },
]


# ----------- Build workbook -----------
def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tab in TABS:
        ws = wb.create_sheet(tab["name"])
        current_row = 1

        for block in tab["blocks"]:
            # Block title (sub-header)
            if block.get("title"):
                title_cell = ws.cell(row=current_row, column=1, value=block["title"])
                title_cell.font = BLOCK_TITLE_FONT
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            # Headers
            for ci, h in enumerate(block["headers"], start=1):
                c = ws.cell(row=current_row, column=ci, value=h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = WRAP
                c.border = BORDER
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            # Data rows
            for row_vals in block["rows"]:
                for ci, val in enumerate(row_vals, start=1):
                    c = ws.cell(row=current_row, column=ci, value=val)
                    c.alignment = WRAP
                    c.border = BORDER
                # Adjust row height based on max content length
                max_len = max((len(str(v)) for v in row_vals), default=0)
                if max_len > 200:
                    ws.row_dimensions[current_row].height = 90
                elif max_len > 100:
                    ws.row_dimensions[current_row].height = 55
                elif max_len > 50:
                    ws.row_dimensions[current_row].height = 36
                else:
                    ws.row_dimensions[current_row].height = 22
                current_row += 1

            # Blank row between blocks
            current_row += 1

        # Column widths
        widths = tab.get("widths", [])
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze top row (header of first block)
        first_block_start = 1
        if tab["blocks"][0].get("title"):
            first_block_start = 2
        # freeze row below the first block's headers
        ws.freeze_panes = ws.cell(row=first_block_start + 1, column=1).coordinate

    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"Arquivo gerado: {path}")
